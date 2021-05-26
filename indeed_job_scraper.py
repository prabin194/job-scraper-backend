from bs4 import BeautifulSoup
import requests
from random import random
from time import sleep
from email.message import EmailMessage
from collections import namedtuple
import smtplib
import csv
from openpyxl import load_workbook

EmailCredentials = namedtuple("EmailCredentials", ['username', 'password', 'sender', 'recipient'])


def generate_url(domain, date_posted, job_title, job_location):
    url_template = "https://" + domain + "/jobs?q={}&l={}&fromage={}"
    url = url_template.format(job_title, job_location, date_posted)
    return url


def save_record_to_csv(record, filepath, create_new_file=False):
    """Save an individual record to file; set `new_file` flag to `True` to generate new file"""
    header = ["JobTitle", "Company", "Location", "Summary", "Salary", "PostDate", "JobUrl"]
    if create_new_file:
        # with open(filepath, mode='w', newline='', encoding='utf-8') as f:
        print(filepath)
        #     # writer = csv.writer(f)
        #     # writer.writerow(header)
        wb = load_workbook(filename=filepath)
        wb.remove(wb.worksheets[0])
        wb.create_sheet()
        ws = wb.worksheets[0]
        ws.append(header)
        wb.save(filepath)

    else:
        # with open(filepath, mode='a+', newline='', encoding='utf-8') as f:
        # writer = csv.writer(f)
        # writer.writerow(record)
        wb = load_workbook(filename=filepath)
        # Select First Worksheet
        ws = wb.worksheets[0]
        ws.append(record)
        wb.save(filepath)


def email_jobs_file(filepath, email):
    """This is currently setup for GMAIL. However, you may need to enable `less secure apps` for
    your email account if you want this to work. See: https://support.google.com/accounts/answer/6010255?hl=en"""
    smtp_host = 'smtp.gmail.com'
    smtp_port = 587
    with smtplib.SMTP(host=smtp_host, port=smtp_port) as server:
        server.starttls()
        server.login(email.username, email.password)
        message = EmailMessage()
        message['From'] = email.sender
        message['To'] = email.recipient
        message['Subject'] = "Updated jobs file"
        message['Body'] = "The updated Indeed postings are attached."
        message.add_attachment(open(filepath, 'r').read(), filename="indeed.csv")
        server.send_message(message)


def collect_job_cards_from_page(html):
    soup = BeautifulSoup(html, 'html.parser')
    cards = soup.find_all('div', 'jobsearch-SerpJobCard')
    return cards, soup


def sleep_for_random_interval():
    seconds = random() * 10
    sleep(seconds)


def request_jobs_from_indeed(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.110 Safari/537.36",
        "Accept-Encoding": "gzip, deflate",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "en"
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.text
    else:
        return None


def find_next_page(soup):
    try:
        pagination = soup.find("a", {"aria-label": "Next"}).get("href")
        return "https://ae.indeed.com" + pagination
    except AttributeError:
        return None


def extract_job_card_data(card):
    atag = card.h2.a
    try:
        job_title = atag.get('title')
    except AttributeError:
        job_title = ''
    try:
        company = card.find('span', 'company').text.strip()
    except AttributeError:
        company = ''
    try:
        location = card.find('div', 'recJobLoc').get('data-rc-loc')
    except AttributeError:
        location = ''
    try:
        job_summary = card.find('div', 'summary').text.strip()
    except AttributeError:
        job_summary = ''
    try:
        post_date = card.find('span', 'date').text.strip()
    except AttributeError:
        post_date = ''
    try:
        salary = card.find('span', 'salarytext').text.strip()
    except AttributeError:
        salary = ''
    job_url = 'https://ae.indeed.com' + atag.get('href')
    return job_title, company, location, job_summary, salary, post_date, job_url


def main(domain, date_posted, job_title, job_location, filepath, email=None):
    unique_jobs = set()  # track job urls to avoid collecting duplicate records
    print("Starting to scrape indeed for `{}` in `{}`".format(job_title, job_location))
    url = generate_url(domain, date_posted, job_title, job_location)
    save_record_to_csv(None, filepath, create_new_file=True)
    page=1

    while page<5:
        html = request_jobs_from_indeed(url)
        if not html:
            break
        cards, soup = collect_job_cards_from_page(html)
        for card in cards:
            record = extract_job_card_data(card)
            if not record[-1] in unique_jobs:
                save_record_to_csv(record, filepath)
                unique_jobs.add(record[-1])
        sleep_for_random_interval()
        url = find_next_page(soup)
        page = page + 1
        if not url:
            break
    print('Finished collecting {:,d} job postings.'.format(len(unique_jobs)))
    if email:
        email_jobs_file(filepath, email)


if __name__ == '__main__':
    # job search settings
    title = 'office assistant'
    loc = ''
    path = 'results.xlsx'

    # include email settings if you want to email the file
    # currently setup for GMAIL... see notes above.
    email_settings = EmailCredentials(
        username='email@gmail.com',
        password='password',
        sender='from@gmail.com',
        recipient='to@gmail.com'
    )

    # using email settings
    # main(title, loc, path, email_settings)

    # without email settings
    main(title, loc, path)
